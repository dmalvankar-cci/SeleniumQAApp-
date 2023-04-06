from openpyxl.reader.excel import load_workbook




def read_data(rowNum, colNum):
        wrkBk = load_workbook("automation_username_password.xlsx")
        sheet = wrkBk.get_sheet_by_name("Sheet1")
        return sheet.cell(row=rowNum, column=colNum).value

def read_data_for_contact_form(rowNum, colNum):
        wrkBk = load_workbook("contact_form_data.xlsx")
        sheet = wrkBk.get_sheet_by_name("Sheet1")
        return sheet.cell(row=rowNum, column=colNum).value

